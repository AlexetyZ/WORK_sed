import pymorphy2

#! / usr / bin / python
# - * - кодировка: latin-1 - * -
import os, sys
import pymorphy2.analyzer
import timestring
from pymorphy2 import opencorpora_dict
import locale
import re
import random



locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
import glob
from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from openpyxl import Workbook
import datetime
import os
from selenium.webdriver.firefox.options import Options
from docxtpl import DocxTemplate
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from docx2pdf import convert
import datetime
from datetime import datetime, date, time, timedelta
import datetime as DT
import time
import timestring
import locale
import pymorphy2
import pymorphy2.analyzer
locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')

from openpyxl.styles import Color, PatternFill, Font, Border
import jinja2
from jinja2 import Template
from collections import OrderedDict

class Workbaze:
    def __init__(self, wb_general='', GP1='', GP0='', chiproc='', fio='', dolzhn='', organization='', subject_of_inspection='', actual_adress='', directly='', when_pass_opros='', vidan_pass_opros='', nomer_pass_opros='', seria_pass_opros='', educ_opros='', native_lang_opros='', dolzhn_controlled='', nation_opros='', fio_controlled='', birthday_opros='', birth_place_opros='', resident_adr_opros='', tel_opros='',  sut='', statia='', srok_protokola='', Dolzhn_spec='', fio_spec='', INN='', OGRN='', Adr_org='', protocol_number='', objectNar='', datetimeNar='', OsnOtv=''):
        self.wb_general = wb_general
        self.GP1 = GP1
        self.GP0 = GP0
        self.chiproc = chiproc


        self.fio = fio

        self.dolzhn = dolzhn
        self.organization = organization

        self.dolzhn_controlled = dolzhn_controlled
        self.fio_controlled = fio_controlled
        self.birthday_opros = birthday_opros
        self.birth_place_opros = birth_place_opros
        self.resident_adr_opros=resident_adr_opros
        self.tel_opros=tel_opros
        self.nation_opros=nation_opros
        self.native_lang_opros=native_lang_opros
        self.educ_opros=educ_opros
        self.seria_pass_opros=seria_pass_opros
        self.nomer_pass_opros=nomer_pass_opros
        self.vidan_pass_opros=vidan_pass_opros
        self.when_pass_opros=when_pass_opros
        self.directly=directly
        self.actual_adress=actual_adress


        self.subject_of_inspection=subject_of_inspection
        self.dolzhn_controlled=dolzhn_controlled


        self.sut = sut
        self.statia = statia
        self.srok_protokola = srok_protokola
        self.Dolzhn_spec = Dolzhn_spec
        self.fio_spec = fio_spec
        self.INN = INN
        self.OGRN = OGRN
        self.Adr_org = Adr_org
        self.protocol_number = protocol_number
        self.objectNar = objectNar
        self.datetimeNar = datetimeNar
        self.OsnOtv = OsnOtv

    def wordskl(self, word, padezh):
        def sklonenie():
            d = OrderedDict()
            split = str.split(word)
            lenspl = len(split)

            if lenspl > 1:

                # print(split[1])
                for i in range(0, lenspl):
                    morph = pymorphy2.MorphAnalyzer()
                    parse_name = morph.parse(split[i])[0]

                    if parse_name.tag.case == 'nomn':
                        n = parse_name.inflect({padezh})[0]
                    else:
                        n = split[i]
                    d[n] = n
                N = list(d.values())

                data = N[0]
                for f in range(1, lenspl):
                    data = data + " " + N[f]
                return data
            else:
                morph = pymorphy2.MorphAnalyzer()
                parse_name = morph.parse(word)[0]
                n = parse_name.inflect({padezh})[0]
                return n

        if re.findall('"', word):
            skl = re.split('"', word)
            word = skl[0]
            return f'{sklonenie()} "{skl[1]}"'

        else:
            return sklonenie()
    def fioskl(self, fio, padezh):

        split = str.split(fio)

        # name
        lingue_name = split[1]
        morph = pymorphy2.MorphAnalyzer()
        parse_name = morph.parse(lingue_name)[0]
        gender_name = parse_name.tag.gender
        I = str.title((parse_name.inflect({padezh})[0]))



        #family
        iskl = ['Дюма', 'Фима', 'Камыса']

        if gender_name == 'masc' and padezh == 'gent':

            splitF = str.split(fio)[0]
            lenf = int(len((splitF)))

            # Мужские фамилии в родительный падеж
            okonchania = {"ев": "ева", "ов": "ова", "й": "ого", "ь": "я", "я": "и", "а": "ы", "ин": "ина"}

            # склоняемые
            # исключения
            if splitF.find('ой', lenf - 2) == lenf - 2:
                if splitF not in iskl:

                    morph = pymorphy2.MorphAnalyzer()
                    parse_name = morph.parse(splitF)
                    var_parse = len(parse_name)
                    n = 0
                    if var_parse > 1:
                        for r in range(0, len(parse_name) - 1):
                            if parse_name[r].tag.POS == 'ADJF':
                                n = n + 1
                                break
                    else:
                        if parse_name[0].tag.POS == 'ADJF':
                            n = n + 1
                    if n > 0:
                        F = str.title(splitF.replace('ой', 'ого', lenf - 2))
                    else:
                        F = str.title(splitF.replace('ой', 'оя', lenf - 2))
                else:
                    F = str.title(splitF)

            if splitF.find('ий', lenf - 2) == lenf - 2:
                F = str.title(splitF.replace('ий', 'ого', lenf - 2))

            if splitF.find('ь', lenf - 1) == lenf - 1:
                F = str.title(splitF.replace('ь', 'я', lenf - 2))

            if splitF.find('я', lenf - 1) == lenf - 1:
                F = str.title(splitF.replace('я', 'и', lenf - 2))

            if splitF.find('а', lenf - 1) == lenf - 1 and splitF[splitF.find('а', lenf - 1) - 1] != 'и':
                F = str.title(splitF.replace('а', 'ы', lenf - 2))


                # основные
            soglasn = ['б', 'в', 'г', 'д', 'ж', 'з', 'к', 'л', 'м', 'н', 'п', 'р', 'с', 'т', 'у', 'ф', 'х', 'ц', 'ч', 'ш',
                       'щ']
            for s in range(len(soglasn)):
                lens = len(soglasn[s])
                posl = splitF.find(soglasn[s], lenf - lens)
                if posl == lenf - lens:
                    F = str.title(splitF.replace(splitF[posl - 2] + splitF[posl - 1] + soglasn[s],
                                                   splitF[posl - 2] + splitF[posl - 1] + soglasn[s] + 'а', -1))

            # не склоняемые
            slovar = ['е', 'и', 'о', 'у', 'ы', 'э', 'ю', 'уа', 'иа']
            for w in range(len(slovar)):
                if splitF.find(slovar[w], lenf - len(slovar[w])) == lenf - len(slovar[w]):
                    F = str.title(splitF)
        if gender_name == 'masc' and padezh == 'ablt':
            splitF = str.split(fio)[0]
            lenf = int(len((splitF)))

            # Мужские фамилии в творительном падеж
            okonchania = {"ев": "ева", "ов": "ова", "й": "ого", "ь": "я", "я": "и", "а": "ы", "ин": "ина"}

            # склоняемые
            # исключения
            if splitF.find('ой', lenf - 2) == lenf - 2:
                if splitF not in iskl:
                    if lenf < 10:
                        morph = pymorphy2.MorphAnalyzer()
                        parse_name = morph.parse(splitF)
                        var_parse = len(parse_name)
                        n = 0
                        if var_parse > 1:
                            for r in range(0, len(parse_name) - 1):
                                if parse_name[r].tag.POS == 'ADJF':
                                    n = n + 1
                                    break
                        else:
                            if parse_name[0].tag.POS == 'ADJF':
                                n = n + 1
                        if n > 0:
                            F = str.title(splitF.replace('ой', 'ым', lenf - 2))
                        else:
                            F = str.title(splitF.replace('ой', 'ойем', lenf - 2))
                else:
                    F = str.title(splitF)
            if splitF.find('ий', lenf - 2) == lenf - 2:
                if splitF not in iskl:
                    F = str.title(
                        splitF.replace(
                            splitF[splitF.find('ий', lenf - 1) - 5] + splitF[splitF.find('ий', lenf - 1) - 4] + splitF[
                                splitF.find('ий', lenf - 1) - 3] + splitF[splitF.find('ий', lenf - 1) - 2] + splitF[
                                splitF.find('ий', lenf - 1) - 1] + 'ий',
                            splitF[splitF.find('ий', lenf - 1) - 5] + splitF[splitF.find('ий', lenf - 1) - 4] + splitF[
                                splitF.find('ий', lenf - 1) - 3] + splitF[splitF.find('ий', lenf - 1) - 2] + splitF[
                                splitF.find('ий', lenf - 1) - 1] + 'им',
                            -1))
                else:
                    F = str.title(splitF)
            if splitF.find('ей', lenf - 2) == lenf - 2:
                if splitF not in iskl:
                    F = str.title(
                        splitF.replace(
                            splitF[splitF.find('ей', lenf - 1) - 5] + splitF[splitF.find('ей', lenf - 1) - 4] + splitF[
                                splitF.find('ей', lenf - 1) - 3] + splitF[splitF.find('йе', lenf - 1) - 2] + splitF[
                                splitF.find('ей', lenf - 1) - 1] + 'йе',
                            splitF[splitF.find('ей', lenf - 1) - 5] + splitF[splitF.find('ей', lenf - 1) - 4] + splitF[
                                splitF.find('ей', lenf - 1) - 3] + splitF[splitF.find('ей', lenf - 1) - 2] + splitF[
                                splitF.find('ей', lenf - 1) - 1] + 'еем',
                            -1))
                else:
                    F = str.title(splitF)
            if splitF.find('ый', lenf - 2) == lenf - 2:
                if splitF not in iskl:
                    F = str.title(
                        splitF.replace(
                            splitF[splitF.find('ый', lenf - 1) - 5] + splitF[splitF.find('ый', lenf - 1) - 4] + splitF[
                                splitF.find('ый', lenf - 1) - 3] + splitF[splitF.find('ый', lenf - 1) - 2] + splitF[
                                splitF.find('ый', lenf - 1) - 1] + 'ый',
                            splitF[splitF.find('ый', lenf - 1) - 5] + splitF[splitF.find('ый', lenf - 1) - 4] + splitF[
                                splitF.find('ый', lenf - 1) - 3] + splitF[splitF.find('ый', lenf - 1) - 2] + splitF[
                                splitF.find('ый', lenf - 1) - 1] + 'ым',
                            -1))
                else:
                    F = str.title(splitF)
            if splitF.find('ь', lenf - 1) == lenf - 1:
                if splitF not in iskl:
                    F = str.title(
                        splitF.replace(
                            splitF[splitF.find('ь', lenf - 1) - 5] + splitF[splitF.find('ь', lenf - 1) - 4] + splitF[
                                splitF.find('ь', lenf - 1) - 3] + splitF[splitF.find('ь', lenf - 1) - 2] + splitF[
                                splitF.find('ь', lenf - 1) - 1] + 'ь',
                            splitF[splitF.find('ь', lenf - 1) - 5] + splitF[splitF.find('ь', lenf - 1) - 4] + splitF[
                                splitF.find('ь', lenf - 1) - 3] + splitF[splitF.find('ь', lenf - 1) - 2] + splitF[
                                splitF.find('ь', lenf - 1) - 1] + 'ем',
                            -1))
                else:
                    F = str.title(splitF)
            if splitF.find('я', lenf - 1) == lenf - 1:
                if splitF not in iskl:
                    F = str.title(
                        splitF.replace(
                            splitF[splitF.find('я', lenf - 1) - 5] + splitF[splitF.find('я', lenf - 1) - 4] + splitF[
                                splitF.find('я', lenf - 1) - 3] + splitF[splitF.find('я', lenf - 1) - 2] + splitF[
                                splitF.find('я', lenf - 1) - 1] + 'я',
                            splitF[splitF.find('я', lenf - 1) - 5] + splitF[splitF.find('я', lenf - 1) - 4] + splitF[
                                splitF.find('я', lenf - 1) - 3] + splitF[splitF.find('я', lenf - 1) - 2] + splitF[
                                splitF.find('я', lenf - 1) - 1] + 'ей',
                            -1))
                else:
                    F = str.title(splitF)

            if splitF.find('а', lenf - 1) == lenf - 1:

                if splitF not in iskl and splitF[splitF.find('а', lenf - 1) - 1] != 'и':
                    F = str.title(
                        splitF.replace(
                            splitF[splitF.find('а', lenf - 1) - 5] + splitF[splitF.find('а', lenf - 1) - 4] + splitF[
                                splitF.find('а', lenf - 1) - 3] + splitF[splitF.find('а', lenf - 1) - 2] + splitF[
                                splitF.find('а', lenf - 1) - 1] + 'а',
                            splitF[splitF.find('а', lenf - 1) - 5] + splitF[splitF.find('а', lenf - 1) - 4] + splitF[
                                splitF.find('а', lenf - 1) - 3] + splitF[splitF.find('а', lenf - 1) - 2] + splitF[
                                splitF.find('а', lenf - 1) - 1] + 'ой',
                            -1))
                else:
                    F = str.title(splitF)
                # основные

            soglasn = ['н', 'в']
            soglasn_em = ['ш', 'щ']
            soglasn_om = ['б', 'г', 'д', 'ж', 'з', 'к', 'л', 'м', 'п', 'р', 'с', 'т', 'ф', 'х', 'ц', 'ч']

            for s in range(len(soglasn)):
                lens = len(soglasn[s])
                posl = splitF.find(soglasn[s], lenf - lens)
                if posl == lenf - lens:
                    F = str.title(splitF.replace(
                        splitF[posl - 4] + splitF[posl - 3] + splitF[posl - 2] + splitF[posl - 1] + soglasn[s],
                        splitF[posl - 4] + splitF[posl - 3] + splitF[posl - 2] + splitF[posl - 1] + soglasn[s] + 'ым', -1))
            for s in range(len(soglasn_em)):
                lens = len(soglasn_em[s])
                posl = splitF.find(soglasn_em[s], lenf - lens)
                if posl == lenf - lens:
                    F = str.title(splitF.replace(
                        splitF[posl - 4] + splitF[posl - 3] + splitF[posl - 2] + splitF[posl - 1] + soglasn_em[s],
                        splitF[posl - 4] + splitF[posl - 3] + splitF[posl - 2] + splitF[posl - 1] + soglasn_em[s] + 'ем',
                        -1))
            for s in range(len(soglasn_om)):
                lens = len(soglasn_om[s])
                posl = splitF.find(soglasn_om[s], lenf - lens)
                if posl == lenf - lens:
                    F = str.title(splitF.replace(
                        splitF[posl - 4] + splitF[posl - 3] + splitF[posl - 2] + splitF[posl - 1] + soglasn_om[s],
                        splitF[posl - 4] + splitF[posl - 3] + splitF[posl - 2] + splitF[posl - 1] + soglasn_om[s] + 'ом',
                        -1))

            # не склоняемые
            slovar = ['е', 'и', 'о', 'у', 'ы', 'э', 'ю', 'уа', 'иа']
            for w in range(len(slovar)):
                if splitF.find(slovar[w], lenf - len(slovar[w])) == lenf - len(slovar[w]):
                    F = str.title(splitF)
        if gender_name == 'femn' and padezh == 'gent':

            splitF = str.split(fio)[0]
            lenf = int(len((splitF)))

            # склоняемые
            # исключения

            if splitF.find('я', lenf - 1) == lenf - 1 and splitF[splitF.find('я', lenf - 1) - 1] != 'и':
                F = str.title(
                        splitF.replace(
                            splitF[splitF.find('ая', lenf - 1) - 5] + splitF[splitF.find('ая', lenf - 1) - 4] + splitF[
                                splitF.find('ая', lenf - 1) - 3] + splitF[splitF.find('ая', lenf - 1) - 2] + splitF[
                                splitF.find('ая', lenf - 1) - 1] + 'ая',
                            splitF[splitF.find('ая', lenf - 1) - 5] + splitF[splitF.find('ая', lenf - 1) - 4] + splitF[
                                splitF.find('ая', lenf - 1) - 3] + splitF[splitF.find('ая', lenf - 1) - 2] + splitF[
                                splitF.find('ая', lenf - 1) - 1] + 'ой',
                            -1))

            if splitF.find('а', lenf - 1) == lenf - 1 and splitF[splitF.find('а', lenf - 1) - 1] != 'и':
                F = str.title(
                        splitF.replace(
                            splitF[splitF.find('а', lenf - 1) - 5] + splitF[splitF.find('а', lenf - 1) - 4] + splitF[
                                splitF.find('а', lenf - 1) - 3] + splitF[splitF.find('а', lenf - 1) - 2] + splitF[
                                splitF.find('а', lenf - 1) - 1] + 'а',
                            splitF[splitF.find('а', lenf - 1) - 5] + splitF[splitF.find('а', lenf - 1) - 4] + splitF[
                                splitF.find('а', lenf - 1) - 3] + splitF[splitF.find('а', lenf - 1) - 2] + splitF[
                                splitF.find('а', lenf - 1) - 1] + 'ой',
                            -1))

            # не склоняемые
            slovar = ['е', 'й', 'и', 'о', 'у', 'ы', 'э', 'ю', 'уа', 'иа', 'ия', 'б', 'в', 'г', 'д', 'ж', 'з', 'к', 'л', 'м',
                      'н', 'п', 'р', 'с', 'т', 'у', 'ф', 'х', 'ц', 'ч', 'ш',
                      'щ']
            for w in range(len(slovar)):
                if splitF.find(slovar[w], lenf - len(slovar[w])) == lenf - len(slovar[w]):
                    F = str.title(splitF)
        if gender_name == 'femn' and padezh == 'ablt':

            splitF = str.split(fio)[0]
            lenf = int(len((splitF)))

            # склоняемые
            # исключения

            if splitF.find('ая', lenf - 1) == lenf - 1 and splitF[splitF.find('я', lenf - 1) - 1] != 'и':
                F = str.title(
                        splitF.replace(
                            splitF[splitF.find('ая', lenf - 1) - 5] + splitF[splitF.find('ая', lenf - 1) - 4] + splitF[
                                splitF.find('ая', lenf - 1) - 3] + splitF[splitF.find('ая', lenf - 1) - 2] + splitF[
                                splitF.find('ая', lenf - 1) - 1] + 'ая',
                            splitF[splitF.find('ая', lenf - 1) - 5] + splitF[splitF.find('ая', lenf - 1) - 4] + splitF[
                                splitF.find('ая', lenf - 1) - 3] + splitF[splitF.find('ая', lenf - 1) - 2] + splitF[
                                splitF.find('ая', lenf - 1) - 1] + 'ой',
                            -1))

            if splitF.find('а', lenf - 1) == lenf - 1 and splitF[splitF.find('а', lenf - 1) - 1] != 'и':
                F = str.title(
                        splitF.replace(
                            splitF[splitF.find('а', lenf - 1) - 5] + splitF[splitF.find('а', lenf - 1) - 4] + splitF[
                                splitF.find('а', lenf - 1) - 3] + splitF[splitF.find('а', lenf - 1) - 2] + splitF[
                                splitF.find('а', lenf - 1) - 1] + 'а',
                            splitF[splitF.find('а', lenf - 1) - 5] + splitF[splitF.find('а', lenf - 1) - 4] + splitF[
                                splitF.find('а', lenf - 1) - 3] + splitF[splitF.find('а', lenf - 1) - 2] + splitF[
                                splitF.find('а', lenf - 1) - 1] + 'ой',
                            -1))

            # не склоняемые
            slovar = ['е', 'й', 'и', 'о', 'у', 'ы', 'э', 'ю', 'уа', 'иа', 'ия', 'б', 'в', 'г', 'д', 'ж', 'з', 'к', 'л', 'м',
                      'н', 'п', 'р', 'с', 'т', 'у', 'ф', 'х', 'ц', 'ч', 'ш',
                      'щ']
            for w in range(len(slovar)):
                if splitF.find(slovar[w], lenf - len(slovar[w])) == lenf - len(slovar[w]):
                    F = str.title(splitF)


        if len(split) > 3:
            O = str.title(split[2]) + ' ' + split[3]
        else:
            lingue = split[2]
            morph = pymorphy2.MorphAnalyzer()
            O = str.title((morph.parse(lingue)[0].inflect({padezh})[0]))

        return (f'{F} {I} {O}')
    def orgskl(self, padezh):
        return self.wordskl(re.split('"', self.organization)[0], padezh) + ' "' + re.split('"', self.organization)[1] + '"'

    def Noweekend(self, date, day):
        for i in range(0, 20, 1):

            dayw = datetime.strptime(date, '%d.%m.%Y') + timedelta(days=day + i)
            dayw_3 = dayw.strftime('%w')

            if int(dayw_3) == 6 or int(dayw_3) == 0:

                continue
            else:

                return dayw
            break
    def Vizov(self):
        # начало автоматического создания документа

        # сейчас

        self.now = datetime.strptime(self.GP0['B7'].value, '%d.%m.%Y %H:%M')
        self.monthfull = str.lower(self.now.strftime('%B'))
        self.month00 = self.now.strftime('%m')
        self.day = self.now.strftime('%d')
        self.year = self.now.strftime('%Y')

        self.fio_gent = self.fioskl(self.fio, 'gent')
        print(f'Родительный падеж (кого?/чего?) - {self.fio_gent}')
        self.fio_ablt = self.fioskl(self.fio, 'ablt')
        print(f'Творительный падеж (кем?/чем?) - {self.fio_ablt}')


        self.dolzhn_ablt = self.wordskl(self.dolzhn, 'ablt')
        self.dolzhn_gent = self.wordskl(self.dolzhn, 'gent')
        self.organization_gent = self.wordskl(re.split('"', self.organization)[0], 'gent')+' "'+re.split('"', self.organization)[1]+'"'


        # в срок через 1 день



        self.dt = datetime.strptime(self.GP0['B7'].value, '%d.%m.%Y %H:%M')
        self.nowa = self.dt.strftime('%d.%m.%Y')
        self.srok = self.Noweekend(self.nowa, self.srok_protokola)
        self.srok_monthfull = str.lower(self.srok.strftime('%B'))
        srok_month00 = self.srok.strftime('%m')
        self.srok_day = self.srok.strftime('%d')
        self.srok_year = self.srok.strftime('%Y')

        self.jinja_env = jinja2.Environment()

        self.doc = DocxTemplate("C:\\Users\\user\\Documents\\документы\\ВЫЗОВ ФОРМА.docx")
        self.doc.render({'день_составления_вызова': self.day, 'месяц_составления_вызова': self.wordskl(str(self.monthfull), 'gent'), 'год_составления_вызова': self.year,
          'Должность_специалиста': self.Dolzhn_spec, 'ФИО_специалиста_полностью': self.fio_spec,
          'Должность_правонарушителя_ТП': self.dolzhn_ablt,
          'организация_правонарушитель': self.orgskl('nomn'), 'организация_правонарушитель_РП': self.orgskl('gent'), 'Суть_правонарушения_что_было_сделано': self.sut,
          'ФИО_правонарушителя_ТП': self.fio_ablt, 'Часть_статья': self.statia, 'Должность_правонарушителя_РП': self.dolzhn_gent, 'ФИО_правонарушителя_РП': self.fio_gent, 'День_составления_протокола': self.srok_day,
          'месяц_составления_протокола': self.wordskl(self.srok_monthfull, 'gent'), 'год_составления_протокола': self.srok_year,
          'ИОФ_специалиста': self.KratkSpec(self.fio_spec)}, self.jinja_env)

        self.doc.save("C:\\Users\\user\\Downloads\\Вызов на протокол " + str.title(self.fio) + ".doc")

    def SvedNar(self): # записывает в exel щаблон на сведения о нарушителе




        print(self.fio)
        if os.path.exists('C:\\Users\\user\\PycharmProjects\\pythonProject\\WorkBaze\\Projects'+'\\'+self.fio+'.xlsx'):

                return 'HUIG'
        else:
            NarInf = Workbook()
            sh1NarInf = NarInf.active
            sh1NarInf.title = 'Информация о нарушителе'
            sh1NarInf.cell(row=2, column=1, value='номер_протокола')
            sh1NarInf.cell(row=2, column=2, value=self.protocol_number)
            sh1NarInf.cell(row=3, column=1, value='выявлено нарушение в ходе проверки(планова/выездная)')
            sh1NarInf.cell(row=3, column=2, value=self.GP0["B5"].value)
            sh1NarInf.cell(row=4, column=1, value='номер_распоряжения_проверки')
            sh1NarInf.cell(row=4, column=2, value=self.GP0["B3"].value)
            sh1NarInf.cell(row=5, column=1, value='дата_распоряжения_проверки')
            sh1NarInf.cell(row=5, column=2, value=self.GP0["B4"].value)
            sh1NarInf.cell(row=6, column=1, value='организация_правонарушитель_полностью')
            sh1NarInf.cell(row=6, column=2, value=self.organization)
            sh1NarInf.cell(row=7, column=1, value='организация_правонарушитель_кратко')
            sh1NarInf.cell(row=7, column=2, value=self.KratkOrg())
            sh1NarInf.cell(row=8, column=1, value='ИНН_организации')
            sh1NarInf.cell(row=8, column=2, value=self.INN)
            sh1NarInf.cell(row=9, column=1, value='ОГРН_организации')
            sh1NarInf.cell(row=9, column=2, value=self.OGRN)
            sh1NarInf.cell(row=10, column=1, value='юр_адрес_организации')
            sh1NarInf.cell(row=10, column=2, value=self.Adr_org)
            sh1NarInf.cell(row=11, column=1, value='адреса_деятельности_организации')
            sh1NarInf.cell(row=11, column=2, value=self.GP0["B14"].value)
            sh1NarInf.cell(row=12, column=1, value='номер_акта_проверки')
            sh1NarInf.cell(row=12, column=2, value=self.GP0["B3"].value)
            sh1NarInf.cell(row=13, column=1, value='дата_акта_проверки')
            sh1NarInf.cell(row=13, column=2, value=self.GP0["B7"].value)
            sh1NarInf.cell(row=14, column=1, value='должность подписавшего распоряжение')
            sh1NarInf.cell(row=14, column=2, value=self.GP0["B9"].value)
            sh1NarInf.cell(row=15, column=1, value='ФИО подписавшего распоряжение')
            sh1NarInf.cell(row=15, column=2, value=self.GP0["B8"].value)
            sh1NarInf.cell(row=16, column=1, value='начало_проверки')
            sh1NarInf.cell(row=16, column=2, value=self.DATA(self.GP0["B6"].value))
            sh1NarInf.cell(row=17, column=1, value='окончание проверки')
            sh1NarInf.cell(row=17, column=2, value=self.DATA(self.GP0["B7"].value))
            sh1NarInf.cell(row=18, column=1, value='Суть_правонарушения_что_было_сделано')
            sh1NarInf.cell(row=18, column=2, value=self.sut)
            sh1NarInf.cell(row=19, column=1, value='должность_правонарушителя')
            sh1NarInf.cell(row=19, column=2, value=self.dolzhn)
            sh1NarInf.cell(row=20, column=1, value='ФИО_правонарушителя_ИП')
            sh1NarInf.cell(row=20, column=2, value=self.fio)
            sh1NarInf.cell(row=21, column=1, value='дата_рождения_правонарушителя')
            # sh1NarInf.cell(row=21, column=2, value=input(f'{self.fio} дата рождения:'))
            sh1NarInf.cell(row=22, column=1, value='серия_номер_паспорта_правонарушителя_кем выдан и когда')
            # seria = input(f'{self.fio} серия паспорта')
            # nomer = input(f'{self.fio} номер паспорта')
            # vidan = input(f'{self.fio} кем выдан паспорт')
            # kogda = input(f'{self.fio} когда выдан паспорт')
            # PASSPORT = f'серия {seria} номер {nomer} выдан {vidan} {kogda}'
            # sh1NarInf.cell(row=22, column=2, value=PASSPORT)
            sh1NarInf.cell(row=23, column=1, value='адрес_места_регистрации_правонарушителя')
            # ADRESS_prop = input(f'{self.fio} адрес_места_регистрации_правонарушителя')
            # sh1NarInf.cell(row=23, column=2, value=ADRESS_prop)
            sh1NarInf.cell(row=24, column=1, value='адрес_места_проживания_правонарушителя')
            # self.ADRESS_prozh = input(f'{self.fio} адрес_места_проживания_правонарушителя если совпадает, просто нажми "ENTER"')
            # if self.ADRESS_prozh == '':
            #     self.ADRESS_prozh = ADRESS_prop
            # sh1NarInf.cell(row=24, column=2, value=self.ADRESS_prozh)
            sh1NarInf.cell(row=25, column=1, value='разговорный_язык_правонарушителя')
            # Langu = input('Оставить разговорный язык правонарушителя по умолчанию РУССКИЙ? если да, просто нажми "ENTER"')
            # if Langu == '':
            #     Langu = 'русский'
            # sh1NarInf.cell(row=25, column=2, value=Langu)
            sh1NarInf.cell(row=26, column=1, value='потребность_в_переводчике')
            # Transl = input('Оставить потребность в переводчике НЕТ по умолчанию?')
            # if Transl == '':
            #     Transl = 'нет'
            # sh1NarInf.cell(row=26, column=2, value=Transl)
            sh1NarInf.cell(row=27, column=1, value='размер_зарплаты_правонарушителя')
            # sh1NarInf.cell(row=27, column=2, value=input(f'{self.fio} размер_зарплаты_правонарушителя'))
            sh1NarInf.cell(row=28, column=1, value='семейное_положение_правонарушителя')
            # sh1NarInf.cell(row=28, column=2, value=input(f'{self.fio} семейное_положение_правонарушителя'))
            sh1NarInf.cell(row=29, column=1, value='всего_иждевенцев_правонарушителя')
            # sh1NarInf.cell(row=29, column=2, value=input(f'{self.fio} всего_иждевенцев_правонарушителя'))
            sh1NarInf.cell(row=30, column=1, value='малолетних_иждевенцев_правонарушителя')
            # sh1NarInf.cell(row=30, column=2, value=input(f'{self.fio} малолетних_иждевенцев_правонарушителя'))
            sh1NarInf.cell(row=31, column=1, value='предыдшущая_административная ответственность (да/нет)')
            # ADMOTVET = input('Оствить Предыдущую административную ответственность НЕТ по умолчанию?')
            # if ADMOTVET == '':
            #     ADMOTVET = 'нет'
            # sh1NarInf.cell(row=31, column=2, value=ADMOTVET)
            sh1NarInf.cell(row=32, column=1, value='часть_статья')
            # sh1NarInf.cell(row=32, column=2, value=self.statia)
            sh1NarInf.cell(row=33, column=1, value='Должность_специалиста_составившего_протокол')
            sh1NarInf.cell(row=33, column=2, value=self.Dolzhn_spec)
            sh1NarInf.cell(row=34, column=1, value='ИОФ_специалиста')
            sh1NarInf.cell(row=34, column=2, value=self.fio_spec)
            sh1NarInf.cell(row=35, column=1, value='место_рождения_правонарушителя')
            # sh1NarInf.cell(row=35, column=2, value=input(f'{self.fio} место_рождения_правонарушителя'))
            sh1NarInf.cell(row=36, column=1, value='СНИЛС_правонарушителя')
            # sh1NarInf.cell(row=36, column=2, value=input(f'{self.fio} СНИЛС_правонарушителя'))




        NarInf.save('C:\\Users\\user\\PycharmProjects\\pythonProject\\WorkBaze\\Projects'+'\\'+self.fio+'.xlsx')


    def KratkOrg(self): # переводит полное название организации в краткую форму путем сокращения букв перед ". например, государственное бюджетное учреждение здравоохранения "Городская больница №37" --> ГБУЗНО "Городская больница №37"
        PATTERN = re.split('"', self.organization)
        pred = ''
        for i in range(0, len(str.split(PATTERN[0]))):
            if len(str.title(str.split(PATTERN[0])[i]))>3:
                if str.title(str.split(PATTERN[0])[i])=='Нижегородской':
                    pred = pred + ' '
                pred = pred+str.title(str.split(PATTERN[0])[i])[0]
        return f'{pred} "{PATTERN[1]}"'


    def KratkSpec(self, fio_spec): # переводит полное ФИО в краткое как для должностного лица. например, Каренина Анна Петровна --> А.П.Каренина
        PATTERN = str.split(fio_spec)
        IOF_spec = f'{str.title(PATTERN[1][0])}.{str.title(PATTERN[2][0])}.{str.title(PATTERN[0])}'

        return IOF_spec
    def DATA(self, datatext): # берет дату формата '%d.%m.%Y %H:%M' и переводит в формат 12:00 01 Мая 2021. Необходимо для заполнения таких сведений как начало проверки.
        datas = datetime.strptime(str(datatext), '%d.%m.%Y %H:%M').strftime('%H:%M %d %B %Y')
        rosn = str.split(datas)

        month = rosn[2]
        morph = pymorphy2.MorphAnalyzer()
        parse_month = morph.parse(month)[0]
        rp_month = str.title((parse_month.inflect({'gent'})[0]))


        return f'{rosn[0]} {rosn[1]} {rp_month} {rosn[3]}'
    def Protocol(self): # формирует протокол об административном правонарушении
        self.SvedNar()
        WBp = openpyxl.load_workbook('C:\\Users\\user\\PycharmProjects\\pythonProject\\WorkBaze\\Projects'+'\\'+self.fio+'.xlsx')
        NARsh = WBp.active


        self.dt = datetime.strptime(self.GP0['B7'].value, '%d.%m.%Y %H:%M')
        d = self.dt.strftime('%d.%m.%Y')
        self.now = self.Noweekend(d, self.srok_protokola)
        self.monthfull = str.lower(self.now.strftime('%B'))
        self.month00 = self.now.strftime('%m')
        self.day = self.now.strftime('%d')
        self.year = self.now.strftime('%Y')

        ruc = self.GP0['B8'].value
        print(self.GP0['B8'].value)
        print(f'это фио руководителя: {ruc}')

        jinja_env = jinja2.Environment()

        doc = DocxTemplate("C:\\Users\\user\\Documents\\документы\\бланки документов\\ПРОТОКОЛ ФОРМА.docx")
        doc.render({'номер_протокола': self.protocol_number, 'день_составления_протокола': self.day, 'месяц_составления_протокола': self.wordskl(self.monthfull, 'gent'),
                    'год_составления_протокола': self.year, 'Должность_специалиста': NARsh['B33'].value,
                    'ФИО_специалиста_полностью': NARsh['B34'].value,
                    'вид_проверки_РП': self.wordskl(self.GP0["B5"].value, 'gent'), 'организация_правонарушитель_полностью_РП': self.orgskl('gent'),
                    'организация_правонарушитель_кратко': self.KratkOrg(), 'юр_адрес_организации': self.Adr_org, 'номер_распоряжения': self.GP0["B3"].value, 'дата_распоряжения': datetime.strptime(str(self.GP0["B4"].value), '%d.%m.%Y').strftime('%d.%m.%Y'),
                    'должность_подписавшего_распоряжение_РП': self.wordskl(self.GP0["B9"].value, 'gent'), 'ФИО_подписавшего_распоряжение_РП': self.fioskl(ruc, 'gent'),
                    'начала_проверки': datetime.strptime(str(self.datetimeNar), '%d.%m.%Y').strftime('%d.%m.%Y'), 'организация_правонарушитель_полностью_ДП': self.orgskl('loct'),
                    'суть_правонарушения_что_было_сделано': self.sut, 'должность_правонарушителя_ИП': NARsh['B19'].value, 'должность_правонарушителя_РП': self.wordskl(NARsh['B19'].value, 'gent'),
                    'организация_правонарушитель_РП': self.KratkOrg(), 'адрес_объекта': self.objectNar, 'основание_ответственности':  self.OsnOtv, 'ФИО_правонарушителя_РП': self.fioskl(self.fio, 'gent'),
                    'номер_акта_проверки': self.GP0["B3"].value, 'дата_акта_проверки': datetime.strptime(str(self.GP0["B7"].value), '%d.%m.%Y %H:%M').strftime('%d.%m.%Y'), 'ФИО_правонарушителя_ИП': self.fio,
                    'дата_рождения_правонарушителя': datetime.strptime(str(NARsh['B21'].value), '%d.%m.%Y').strftime('%d.%m.%Y'), 'серия_номер_паспорта_правонарушителя': NARsh['B22'].value,
                    'адрес_места_регистрации_правонарушителя': NARsh['B23'].value, 'адрес_места_проживания_правонарушителя': NARsh['B24'].value,
                    'разговорный_язык_правонарушителя': NARsh['B25'].value, 'потребность_в_переводчике': NARsh['B26'].value, 'размер_зарплаты_правонарушителя': NARsh['B27'].value,
                    'семейное_положение_правонарушителя': NARsh['B28'].value, 'всего_иждевенцев_правонарушителя': NARsh['B29'].value,
                    'малолетних_иждевенцев_правонарушителя': NARsh['B30'].value, 'предыдшущая_административная_ответственность': NARsh['B31'].value, 'часть_статья': NARsh['B32'].value,
                    'ИОФ_специалиста': self.KratkSpec(self.fio_spec), 'место_рождения_правонарушителя': NARsh['B35'].value, 'СНИЛС_правонарушителя': NARsh['B36'].value}, jinja_env)

        doc.save("C:\\Users\\user\\Downloads\\Протокол " + str.title(self.fio) + ".doc")

    def Protocol_oprosa(self): # формирует протокол опроса
        now = datetime.now()
        monthfull = str.lower(now.strftime('%B'))
        month00 = now.strftime('%m')
        day = now.strftime('%d')
        year = now.strftime('%Y')
        hour_start = now.strftime('%H')
        minute_start = now.strftime('%M')
        finish = datetime.now()+timedelta(minutes=random.randint(9, 15))

        hour_finish = finish.strftime('%H')
        minute_finish = finish.strftime('%M')

        ps = str(self.GP0["B6"].value)  # proverka starts/ ссылается на ячейку даты начала проверки
        day_ps = datetime.strptime(ps, '%d.%m.%Y %H:%M').strftime('%d')
        month_ps = datetime.strptime(ps, '%d.%m.%Y %H:%M').strftime('%m')
        year_ps = datetime.strptime(ps, '%d.%m.%Y %H:%M').strftime('%Y')

        print(self.fio_controlled)


        jinja_env = jinja2.Environment()

        doc = DocxTemplate("C:\\Users\\user\\PycharmProjects\\pythonProject\\WorkBaze\\БЛАНКИ\\ПРОТОКОЛ ОПРОСА.docx")
        doc.render({'деньсоставления': day, 'месяцcоставления': self.wordskl(monthfull, 'gent'), 'годсоставления': year, 'место_составления': self.actual_adress, 'должность_специалиста_ИП': self.Dolzhn_spec, 'ИОФамилия_специалиста': self.KratkSpec(self.fio_spec), 'час_составления': hour_start, 'минуты_составления': minute_start, 'час_окончания': hour_finish, 'минуты_окончания': minute_finish, 'вид_проверки_РП': self.GP0["B5"].value, 'день_начала_проверки': day_ps, 'месяц_начала_проверки': month_ps, 'год_начала_проверки': year_ps, 'номер_проверки': self.GP0["B3"].value, 'должность_опрашиваемого_РП': self.wordskl(self.dolzhn_controlled, 'gent'), 'ИОФамилия_опрашиваемого': self.KratkSpec(self.fio_controlled), 'ФИО_опрашиваемого': self.fio_controlled, 'дата_рождения_опрашиваемого': self.birthday_opros, 'место_рождения_опрашиваемого': self.birth_place_opros, 'место_жительства_опрашиваемого': self.resident_adr_opros, 'телефон_опрашиваемого': self.tel_opros, 'гражданство_опрашиваемого': self.nation_opros, 'родной_язык_опрашиваемого': self.native_lang_opros, 'образование_опрашиваемого': self.educ_opros, 'место_работы_опрашиваемого': f'{self.KratkOrg()} {self.dolzhn_controlled}', 'должность_опрашиваемого_ИП': self.dolzhn_controlled, 'серия_паспорта_опрашиваемого': self.seria_pass_opros, 'номер_паспорта_опрашиваемого': self.nomer_pass_opros, 'кем_выдан_паспорт_опрашиваемого': self.vidan_pass_opros, 'когда_выдан_паспорт_опрашиваемого': self.when_pass_opros, 'опрос_по_существу': self.directly,}, jinja_env)

        doc.save("C:\\Users\\user\\Downloads\\Протокол опроса " + str.title(self.fio_controlled) + ".doc")

    def Protocol_osmotra(self):
        now = datetime.now()
        monthfull = str.lower(now.strftime('%B'))
        month00 = now.strftime('%m')
        day = now.strftime('%d')
        year = now.strftime('%Y')
        hour_start = now.strftime('%H')
        minute_start = now.strftime('%M')
        finish = datetime.now()+timedelta(minutes=random.randint(50, 120))

        hour_finish = finish.strftime('%H')
        minute_finish = finish.strftime('%M')

        ps = str(self.GP0["B6"].value)  # proverka starts/ ссылается на ячейку даты начала проверки
        day_ps = datetime.strptime(ps, '%d.%m.%Y %H:%M').strftime('%d')
        month_ps = datetime.strptime(ps, '%d.%m.%Y %H:%M').strftime('%m')
        year_ps = datetime.strptime(ps, '%d.%m.%Y %H:%M').strftime('%Y')

        print(self.wordskl(monthfull, 'gent'))




        jinja_env = jinja2.Environment()

        doc = DocxTemplate("C:\\Users\\user\\PycharmProjects\\pythonProject\\WorkBaze\\БЛАНКИ\\ПРОТОКОЛ ОСМОТРА.docx")
        doc.render({'деньсоставления': day,
                    'месяцcоставления': self.wordskl(monthfull, 'gent'),
                    'годсоставления': year,
                    'место_составления': self.GP0["B10"].value[0].upper()+self.wordskl(self.GP0["B10"].value, 'gent')[1:],
                    'должность_специалиста_ИП': f'{self.Dolzhn_spec} Территориального отдела Управления Роспотребнадзора по Нижегородской области в Автозаводском, Ленинском районах г.Нижнего Новгорода и Богородском районе',
                    'ФИО_специалиста': self.fio_spec,
                    'Произведен_осмотр_чего': self.subject_of_inspection,
                    'Организация_кратко': self.KratkOrg(),
                    'Адрес_ораганизации': self.actual_adress,
                    'час_составления': hour_start,
                    'минуты_составления': minute_start,
                    'час_окончания': hour_finish,
                    'минуты_окончания': minute_finish,
                    'Должность_представителя_РП': self.wordskl(self.dolzhn_controlled, 'gent'),
                    'ФИО_представителя_РП': self.fioskl(self.fio_controlled, 'gent'),
                    'осмотр_по_существу': self.directly,
                    'ФИО_представителя_ИП': self.KratkSpec(self.fio_controlled),
                    'ИОФамилия_специалиста': self.KratkSpec(self.fio_spec),}, jinja_env
        )

        doc.save("C:\\Users\\user\\Downloads\\Протокол осмотра " + str.title(self.organization.replace('"', '')) + ".doc")
        doc.save("C:\\Users\\user\\Downloads\\Протокол осмотра " + str.title(self.organization.replace('"', '')) + now.strftime('%H%M%S') + ".doc")




    def Opredelenie(self):
        WBp = openpyxl.load_workbook('C:\\Users\\user\\PycharmProjects\\pythonProject\\WorkBaze\\Projects'+'\\'+self.fio+'.xlsx')
        NARsh = WBp.active


        self.dt = datetime.strptime(self.GP0['B7'].value, '%d.%m.%Y %H:%M')
        d = self.dt.strftime('%d.%m.%Y')
        self.now = self.Noweekend(d, self.srok_protokola)
        self.monthfull = str.lower(self.now.strftime('%B'))
        self.month00 = self.now.strftime('%m')
        self.day = self.now.strftime('%d')
        self.year = self.now.strftime('%Y')

        jinja_env = jinja2.Environment()

        doc = DocxTemplate("C:\\Users\\user\\Documents\\документы\\бланки документов\\ОПРЕДЕЛЕНИЕ ФОРМА.docx")
        doc.render({'номер_протокола': self.protocol_number,
                    'день_составления_протокола': self.day,
                    'месяц_составления_протокола': self.wordskl(self.monthfull, 'gent'),
                    'год_составления_протокола': self.year,
                    'ФИО_правонарушителя_РП': self.fioskl(self.fio, 'gent'),
                    'дата_рождения_правонарушителя': datetime.strptime(str(NARsh['B21'].value),
                                                                       '%d.%m.%Y').strftime('%d.%m.%Y'),
                    'место_рождения_правонарушителя': NARsh['B35'].value,
                    'должность_правонарушителя_РП': self.wordskl(NARsh['B19'].value, 'gent'),
                    'организация_правонарушитель_кратко': self.KratkOrg(),
                    'адрес_объекта': self.objectNar,
                    'часть_статья': NARsh['B32'].value,
                        }, jinja_env)

        doc.save("C:\\Users\\user\\Downloads\\Определение " + str.title(self.fio) + ".doc")
