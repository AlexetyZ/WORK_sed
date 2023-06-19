from учз import SedRequests
import openpyxl


def main(path):
    wb = openpyxl.load_workbook(path)
    sh = wb.worksheets[0]
    session = SedRequests(userName='Зайцев А.Д')
    for row in sh.iter_rows(min_row=7, values_only=True):
        if not row[1]:
            break
        number = str(row[1]).strip()
        author = str(row[3]).split('\n')[0].strip()
        print(f'number - {number}')
        session.downloadFiles(number, author)


def test(path):
    wb = openpyxl.load_workbook(path)
    sh = wb.worksheets[0]
    for row in sh.iter_rows(min_row=7, values_only=True):
        if not row[1]:
            break
        who = str(row[3]).split('\n')[0].strip()
        print(who)


if __name__ == '__main__':
    path_file = "S:\\Зайцев_АД\скачанные из сэд\Список_выбранных_документов.xlsx"
    main(path_file)
    # test(path_file)
